# -*- coding: utf-8 -*-
"""
VitaKa - Программа для учета комплектующих на заводе
Версия: 1.0
"""
from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
import os
import sys
import json
import copy
import time
import subprocess

# ─────────────────────────────────────────────────────────────
#  КОНСТАНТЫ
# ─────────────────────────────────────────────────────────────
SETTINGS_FILE = "vitaka_settings.json"
DEFAULT_DB_NAME = "vitaka_components.xlsx"
SHEET_TYPES = "Типы"
SHEET_COMPONENTS = "Комплектующие"
SHEET_LOG = "Лог"

TYPES_COLUMNS = ["id", "название", "описание", "параметры"]
TYPES_HEADERS = {
    "id": "ID",
    "название": "Название типа",
    "описание": "Описание",
    "параметры": "Параметры",
}

# Параметры по умолчанию для каждого типа
DEFAULT_TYPE_PARAMETERS = ["Количество", "Комментарий"]

STANDARD_COLUMNS = ["id", "тип", "диаметр", "длина", "количество", "вес_единицы"]
STANDARD_HEADERS = {
    "id": "ID",
    "тип": "Тип",
    "диаметр": "Диаметр",
    "длина": "Длина (мм)",
    "количество": "Количество",
    "вес_единицы": "Вес/ед. (кг)",
}
LOG_COLUMNS = ["дата_время", "операция", "комплектующее", "изменение", "комментарий"]
LOG_HEADERS = {
    "дата_время": "Дата и время",
    "операция": "Операция",
    "комплектующее": "Комплектующее",
    "изменение": "Изменение",
    "комментарий": "Комментарий",
}

COMPONENT_TYPES = ["Винт", "Болт", "Гайка", "Шайба", "Шпилька", "Заклёпка", "Штифт", "Другое"]
DIAMETER_OPTIONS = ["М2", "М2.5", "М3", "М4", "М5", "М6", "М8", "М10", "М12", "М16", "М20",
                    "3", "4", "5", "6", "8", "9", "10", "12", "16", "20"]


# ─────────────────────────────────────────────────────────────
#  НАСТРОЙКИ
# ─────────────────────────────────────────────────────────────

def load_settings() -> dict:
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_settings(settings: dict):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Ошибка сохранения настроек: {e}")


# ─────────────────────────────────────────────────────────────
#  РАБОТА С EXCEL
# ─────────────────────────────────────────────────────────────

def get_db_path(settings: dict) -> str:
    folder = settings.get("db_folder", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(folder, DEFAULT_DB_NAME)


def initialize_db(file_path: str):
    """Создать файл Excel с нужными листами, если не существует."""
    if os.path.exists(file_path):
        return
    wb = Workbook()

    # Лист типов
    ws_types = wb.active
    ws_types.title = SHEET_TYPES
    ws_types.append(["id", "название", "описание"])

    # Лист комплектующих
    ws_comp = wb.create_sheet(SHEET_COMPONENTS)
    ws_comp.append(["id", "type_id", "тип", "диаметр", "длина", "количество", "вес_единицы", "доп_параметры"])

    # Лист логов
    ws_log = wb.create_sheet(SHEET_LOG)
    ws_log.append(["дата_время", "операция", "комплектующее", "изменение", "комментарий"])

    wb.save(file_path)


def load_components(file_path: str) -> list:
    """Загрузить список комплектующих из Excel."""
    try:
        initialize_db(file_path)
        df = pd.read_excel(file_path, sheet_name=SHEET_COMPONENTS, dtype=str)
        df = df.fillna("")
        items = []
        for _, row in df.iterrows():
            item = {col: row[col] for col in STANDARD_COLUMNS if col in df.columns}
            # Дополнительные параметры хранятся как JSON в колонке "доп_параметры"
            extra_raw = row.get("доп_параметры", "") if "доп_параметры" in df.columns else ""
            try:
                item["доп_параметры"] = json.loads(extra_raw) if extra_raw else {}
            except Exception:
                item["доп_параметры"] = {}
            items.append(item)
        return items
    except Exception as e:
        print(f"Ошибка загрузки комплектующих: {e}")
        return []


def load_log(file_path: str) -> list:
    """Загрузить лог из Excel."""
    try:
        initialize_db(file_path)
        df = pd.read_excel(file_path, sheet_name=SHEET_LOG, dtype=str)
        df = df.fillna("")
        return df.to_dict("records")
    except Exception as e:
        print(f"Ошибка загрузки лога: {e}")
        return []


def load_types(file_path: str) -> list:
    """Загрузить список типов комплектующих из Excel."""
    try:
        initialize_db(file_path)
        # Проверяем наличие листа
        if SHEET_TYPES not in pd.ExcelFile(file_path).sheet_names:
            return []
        df = pd.read_excel(file_path, sheet_name=SHEET_TYPES, dtype=str)
        df = df.fillna("")
        items = []
        for _, row in df.iterrows():
            item = {col: row[col] for col in TYPES_COLUMNS if col in df.columns}
            # Параметры хранятся как JSON
            params_raw = item.get("параметры", "") if "параметры" in df.columns else ""
            try:
                item["параметры"] = json.loads(params_raw) if params_raw else DEFAULT_TYPE_PARAMETERS.copy()
            except Exception:
                item["параметры"] = DEFAULT_TYPE_PARAMETERS.copy()
            items.append(item)
        return items
    except Exception as e:
        print(f"Ошибка загрузки типов: {e}")
        return []


def next_type_id(types: list) -> int:
    if not types:
        return 1
    try:
        return max(int(t.get("id", 0)) for t in types) + 1
    except Exception:
        return len(types) + 1


def save_all_with_types(file_path: str, types: list, components: list, log_entries: list):
    """Сохранить все данные включая типы в Excel."""
    try:
        folder = os.path.dirname(file_path)
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)

        # Типы (преобразуем параметры в JSON)
        types_for_save = []
        for t in types:
            t_copy = dict(t)
            t_copy["параметры"] = json.dumps(t.get("параметры", DEFAULT_TYPE_PARAMETERS), ensure_ascii=False)
            types_for_save.append(t_copy)
        df_types = pd.DataFrame(types_for_save, columns=TYPES_COLUMNS) if types_for_save else pd.DataFrame(
            columns=TYPES_COLUMNS)

        # Комплектующие
        rows = []
        for item in components:
            row = {col: item.get(col, "") for col in STANDARD_COLUMNS if col != "тип"}
            row["type_id"] = item.get("type_id", "")
            row["тип"] = item.get("тип", "")
            row["доп_параметры"] = json.dumps(item.get("доп_параметры", {}), ensure_ascii=False)
            rows.append(row)

        comp_cols = ["id", "type_id"] + STANDARD_COLUMNS + ["доп_параметры"]
        df_comp = pd.DataFrame(rows, columns=comp_cols) if rows else pd.DataFrame(columns=comp_cols)

        df_log = pd.DataFrame(log_entries, columns=LOG_COLUMNS) if log_entries else pd.DataFrame(columns=LOG_COLUMNS)

        # Сохранение
        if os.path.exists(file_path):
            with pd.ExcelFile(file_path, engine="openpyxl") as xls:
                other_sheets = {s: pd.read_excel(xls, s) for s in xls.sheet_names
                                if s not in (SHEET_TYPES, SHEET_COMPONENTS, SHEET_LOG)}
        else:
            other_sheets = {}

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_types.to_excel(writer, sheet_name=SHEET_TYPES, index=False)
            df_comp.to_excel(writer, sheet_name=SHEET_COMPONENTS, index=False)
            df_log.to_excel(writer, sheet_name=SHEET_LOG, index=False)
            for name, df in other_sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)
    except Exception as e:
        print(f"Ошибка сохранения Excel: {e}")
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл:\n{e}")

def save_all(file_path: str, components: list, log_entries: list):
    """Сохранить все данные в Excel, заменяя оба листа."""
    try:
        folder = os.path.dirname(file_path)
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)

        # Подготовить DataFrame для комплектующих
        rows = []
        for item in components:
            row = {col: item.get(col, "") for col in STANDARD_COLUMNS}
            row["доп_параметры"] = json.dumps(item.get("доп_параметры", {}), ensure_ascii=False)
            rows.append(row)
        df_comp = pd.DataFrame(rows, columns=STANDARD_COLUMNS + ["доп_параметры"])

        df_log = pd.DataFrame(log_entries, columns=LOG_COLUMNS) if log_entries else pd.DataFrame(
            columns=LOG_COLUMNS)

        # Если файл существует – сохраняем с остальными листами
        if os.path.exists(file_path):
            with pd.ExcelFile(file_path, engine="openpyxl") as xls:
                other_sheets = {s: pd.read_excel(xls, s) for s in xls.sheet_names
                                if s not in (SHEET_COMPONENTS, SHEET_LOG)}
        else:
            other_sheets = {}

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_comp.to_excel(writer, sheet_name=SHEET_COMPONENTS, index=False)
            df_log.to_excel(writer, sheet_name=SHEET_LOG, index=False)
            for name, df in other_sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)
    except Exception as e:
        print(f"Ошибка сохранения Excel: {e}")
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл:\n{e}")


# ─────────────────────────────────────────────────────────────
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────────────────────────

def next_id(components: list) -> int:
    if not components:
        return 1
    try:
        return max(int(c.get("id", 0)) for c in components) + 1
    except Exception:
        return len(components) + 1


def component_label(item: dict) -> str:
    """Читаемое название комплектующего для лога."""
    parts = []
    if item.get("тип"):
        parts.append(str(item["тип"]))
    if item.get("диаметр"):
        parts.append(str(item["диаметр"]))
    if item.get("длина"):
        parts.append(f"x{item['длина']}")
    return " ".join(parts) if parts else f"ID={item.get('id', '?')}"


def diff_items(old: dict, new: dict) -> str:
    """Описать разницу между двумя версиями комплектующего."""
    changes = []
    fields = {
        "тип": "Тип",
        "диаметр": "Диаметр",
        "длина": "Длина",
        "количество": "Количество",
        "вес_единицы": "Вес/ед.",
    }
    for key, label in fields.items():
        ov = str(old.get(key, "")).strip()
        nv = str(new.get(key, "")).strip()
        if ov != nv:
            changes.append(f"{label}: {ov} → {nv}")

    # Дополнительные параметры
    old_extra = old.get("доп_параметры", {})
    new_extra = new.get("доп_параметры", {})
    all_keys = set(old_extra) | set(new_extra)
    for k in sorted(all_keys):
        ov = str(old_extra.get(k, "")).strip()
        nv = str(new_extra.get(k, "")).strip()
        if ov != nv:
            if not ov:
                changes.append(f"{k}: добавлено «{nv}»")
            elif not nv:
                changes.append(f"{k}: удалено «{ov}»")
            else:
                changes.append(f"{k}: {ov} → {nv}")

    return "; ".join(changes) if changes else "Без изменений"


# ─────────────────────────────────────────────────────────────
#  ДИАЛОГ ДОБАВЛЕНИЯ / РЕДАКТИРОВАНИЯ КОМПЛЕКТУЮЩЕГО
# ─────────────────────────────────────────────────────────────

class ComponentDialog(tk.Toplevel):
    """Диалоговое окно для создания или редактирования комплектующего."""

    def __init__(self, parent, title: str, item: dict = None, app=None):
        super().__init__(parent)
        self.title(title)
        self.resizable(True, True)
        self.grab_set()
        self.result = None

        self._app = app  # 🆕 ДОБАВЬ ЭТУ СТРОКУ!
        self._item = copy.deepcopy(item) if item else {}
        self._extra_rows = []  # список (key_var, val_var, frame)

        self._build_ui()
        self._populate(self._item)

        self.update_idletasks()
        w, h = 480, 520
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(400, 400)

    # ── построение интерфейса ──────────────────────────────

    def _build_ui(self):
        outer = ttk.Frame(self, padding=10)
        outer.pack(fill="both", expand=True)

        # Прокручиваемая область
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._scroll_frame = ttk.Frame(canvas)
        self._canvas_window = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")

        self._scroll_frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._canvas_window, width=e.width))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        form = self._scroll_frame
        r = 0

        # Стандартные поля
        # 🆕 ПАРАМЕТР ТИП (название типа из списка типов)
        ttk.Label(form, text="Тип *").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._type_name_var = tk.StringVar()
        # Получаем список названий типов из app
        type_names = []
        if self._app and hasattr(self._app, 'component_types'):
            type_names = [t.get("название", "") for t in self._app.component_types]
        cb_type_name = ttk.Combobox(form, textvariable=self._type_name_var, values=type_names, width=28)
        cb_type_name.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Диаметр").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._diam_var = tk.StringVar()
        cb_diam = ttk.Combobox(form, textvariable=self._diam_var, values=DIAMETER_OPTIONS, width=28)
        cb_diam.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Длина (мм)").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._len_var = tk.StringVar()
        ttk.Entry(form, textvariable=self._len_var, width=30).grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        # 🆕 НЕ ПОКАЗЫВАЕМ стандартные поля Количество и Вес/ед.
        # Они будут в параметрах типа!

        form.columnconfigure(1, weight=1)

        # Разделитель для параметров
        ttk.Separator(form, orient="horizontal").grid(row=r, column=0, columnspan=2,
                                                      sticky="ew", padx=5, pady=8)
        r += 1
        ttk.Label(form, text="Параметры типа:", font=("", 9, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=5)
        r += 1

        self._extra_container = ttk.Frame(form)
        self._extra_container.grid(row=r, column=0, columnspan=2, sticky="ew", padx=5)
        self._extra_container.columnconfigure(0, weight=1)
        self._extra_container.columnconfigure(1, weight=1)
        r += 1

        # Кнопки OK / Отмена
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", side="bottom", padx=10, pady=8)
        ttk.Button(btn_frame, text="Сохранить", command=self._on_ok, width=14).pack(side="right", padx=4)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy, width=10).pack(side="right", padx=4)

    def _add_extra_row(self, key: str = "", value: str = ""):
        row_frame = ttk.Frame(self._extra_container)
        idx = len(self._extra_rows)
        row_frame.grid(row=idx, column=0, columnspan=3, sticky="ew", pady=2)
        row_frame.columnconfigure(0, weight=1)
        row_frame.columnconfigure(1, weight=1)

        key_var = tk.StringVar(value=key)
        val_var = tk.StringVar(value=value)
        ttk.Entry(row_frame, textvariable=key_var, width=16,
                  placeholder_text="Название").grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Entry(row_frame, textvariable=val_var, width=16).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(row_frame, text="✕", width=3,
                   command=lambda f=row_frame, kv=key_var, vv=val_var: self._remove_extra_row(f, kv, vv)).grid(
            row=0, column=2)
        self._extra_rows.append((key_var, val_var, row_frame))

    def _remove_extra_row(self, frame, key_var, val_var):
        self._extra_rows = [(k, v, f) for k, v, f in self._extra_rows if k is not key_var]
        frame.destroy()

    # ── заполнение данными ─────────────────────────────────

    def _populate(self, item: dict):
        # 🆕 Устанавливаем только название типа (если редактируем)
        if item.get("id"):  # Если редактируем существующее
            self._type_name_var.set(item.get("тип", ""))

            self._diam_var.set(item.get("диаметр", ""))
            self._len_var.set(item.get("длина", ""))

            # ДОБАВЛЯЕМ ПАРАМЕТРЫ ИЗ ТИПА
            type_name = item.get("тип", "")
            if type_name and self._app and hasattr(self._app, 'component_types'):
                type_item = next((t for t in self._app.component_types if t.get("название") == type_name), None)
                if type_item:
                    type_params = type_item.get("параметры", [])
                    for param_name in type_params:
                        param_value = item.get("доп_параметры", {}).get(param_name, "")
                        self._add_extra_row(param_name, param_value)
        else:  # 🆕 Если создаём новое комплектующее
            # При создании параметры добавляются только если тип выбран
            pass

    def _on_ok(self):
        # Получаем название типа из combobox
        selected_type = self._type_name_var.get().strip()
        if not selected_type:
            messagebox.showwarning("Внимание", "Поле «Тип» обязательно для заполнения.", parent=self)
            return

        # Валидация длины
        length_value = self._len_var.get().strip()
        if length_value and not self._is_number(length_value):
            messagebox.showwarning("Внимание", "Длина должна быть числом.", parent=self)
            return

        # Собираем параметры
        params = {}
        for key_var, val_var, _ in self._extra_rows:
            param_key = key_var.get().strip()
            param_val = val_var.get().strip()
            if param_key:
                params[param_key] = param_val

        # Результат
        self.result = {
            "id": self._item.get("id", ""),
            "тип": selected_type,
            "диаметр": self._diam_var.get().strip(),
            "длина": length_value,
            "количество": "",
            "вес_единицы": "",
            "доп_параметры": params,
        }
        self.destroy()

    @staticmethod
    def _is_number(s: str) -> bool:
        try:
            float(s.replace(",", "."))
            return True
        except ValueError:
            return False


class ComponentTypeDialog(tk.Toplevel):
    """Диалоговое окно для создания или редактирования типа комплектующего."""

    def __init__(self, parent, title: str, item: dict = None):
        super().__init__(parent)
        self.title(title)
        self.resizable(True, True)
        self.grab_set()
        self.result = None
        self._app = app  # 🆕 Сохраняем ссылку на приложение

        self._item = copy.deepcopy(item) if item else {}
        self._param_rows = []  # список (param_var, frame)
        self._build_ui()
        self._populate(self._item)

        self.update_idletasks()
        w, h = 500, 450
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(400, 350)

    def _build_ui(self):
        outer = ttk.Frame(self, padding=10)
        outer.pack(fill="both", expand=True)

        # Прокручиваемая область
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._scroll_frame = ttk.Frame(canvas)
        self._canvas_window = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")

        self._scroll_frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._canvas_window, width=e.width))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        form = self._scroll_frame
        r = 0

        # Название
        ttk.Label(form, text="Название *").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._name_var = tk.StringVar()
        ttk.Entry(form, textvariable=self._name_var, width=40).grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        # Описание
        ttk.Label(form, text="Описание").grid(row=r, column=0, sticky="nw", padx=5, pady=4)
        self._desc_text = tk.Text(form, height=3, width=40)
        self._desc_text.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        form.columnconfigure(1, weight=1)

        # Разделитель
        ttk.Separator(form, orient="horizontal").grid(row=r, column=0, columnspan=2,
                                                      sticky="ew", padx=5, pady=8)
        r += 1

        # Параметры
        ttk.Label(form, text="Параметры типа:", font=("", 9, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=5)
        r += 1

        self._params_container = ttk.Frame(form)
        self._params_container.grid(row=r, column=0, columnspan=2, sticky="ew", padx=5)
        self._params_container.columnconfigure(0, weight=1)
        r += 1

        ttk.Button(form, text="➕ Добавить параметр", command=self._add_param_row).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="⚠️ Параметры нельзя будет менять после создания типа!",
                  font=("", 8), foreground="red").grid(row=r, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        r += 1

        # Кнопки OK / Отмена
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", side="bottom", padx=10, pady=8)
        ttk.Button(btn_frame, text="Сохранить", command=self._on_ok, width=14).pack(side="right", padx=4)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy, width=10).pack(side="right", padx=4)

    def _add_param_row(self, param_name: str = ""):
        row_frame = ttk.Frame(self._params_container)
        idx = len(self._param_rows)
        row_frame.grid(row=idx, column=0, sticky="ew", pady=2)
        row_frame.columnconfigure(0, weight=1)

        param_var = tk.StringVar(value=param_name)
        ttk.Entry(row_frame, textvariable=param_var, width=30).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Button(row_frame, text="✕", width=3,
                   command=lambda f=row_frame, pv=param_var: self._remove_param_row(f, pv)).grid(row=0, column=1)

        self._param_rows.append((param_var, row_frame))

    def _remove_param_row(self, frame, param_var):
        self._param_rows = [(pv, f) for pv, f in self._param_rows if pv is not param_var]
        frame.destroy()

    def _populate(self, item: dict):
        self._name_var.set(item.get("название", ""))
        self._desc_text.insert("1.0", item.get("описание", ""))

        # Параметры (если редактируем - показываем что есть, если создаём - добавляем стандартные)
        params = item.get("параметры", DEFAULT_TYPE_PARAMETERS)
        for param in params:
            self._add_param_row(param)

        # Если новый тип - добавляем стандартные параметры
        if not item.get("id"):
            if not self._param_rows:
                for param in DEFAULT_TYPE_PARAMETERS:
                    self._add_param_row(param)

    def _on_ok(self):
        название = self._name_var.get().strip()
        if not название:
            messagebox.showwarning("Внимание", "Поле «Название» обязательно для заполнения.", parent=self)
            return

        описание = self._desc_text.get("1.0", "end").strip()

        # Собираем параметры
        параметры = []
        for param_var, _ in self._param_rows:
            p = param_var.get().strip()
            if p:
                параметры.append(p)

        if not параметры:
            messagebox.showwarning("Внимание", "Добавьте хотя бы один параметр.", parent=self)
            return

        self.result = {
            "id": self._item.get("id", ""),
            "название": название,
            "описание": описание,
            "параметры": параметры,
        }
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  EXCEL-СТИЛЬ ФИЛЬТР (НОВАЯ ВЕРСИЯ)
# ─────────────────────────────────────────────────────────────

class ExcelStyleFilter:
    """Фильтр в стиле Excel для Treeview - выпадающее меню при клике на заголовок"""

    def __init__(self, tree, refresh_callback, columns_config=None):
        """
        tree: ttk.Treeview виджет
        refresh_callback: функция обновления данных
        columns_config: dict с настройками столбцов (опционально)
        """
        self.tree = tree
        self.refresh_callback = refresh_callback
        self.columns_config = columns_config or {}

        # Хранилище активных фильтров
        self.active_filters = {}

        # Исходные данные (до фильтрации)
        self.original_data = []

        # ЗАЩИТА ОТ ДВОЙНОГО ВЫЗОВА
        self._filter_window_open = False
        self._last_click_time = 0

        # Привязываем клик к заголовкам
        self.tree.bind('<Button-1>', self.on_header_click)

    def on_header_click(self, event):
        """Обработка клика по заголовку столбца"""
        import time

        # ЗАЩИТА ОТ ДВОЙНОГО КЛИКА
        current_time = time.time()
        if current_time - self._last_click_time < 0.3:
            return

        # ЗАЩИТА: НЕ ОТКРЫВАЕМ ВТОРОЕ ОКНО
        if self._filter_window_open:
            return

        region = self.tree.identify_region(event.x, event.y)

        if region == "heading":
            column = self.tree.identify_column(event.x)
            column_id = self.tree.column(column, "id")

            self._last_click_time = current_time
            self._filter_window_open = True

            # Показываем меню фильтра
            self.show_filter_menu(event, column_id)

    def reapply_all_filters(self):
        """Переприменить все активные фильтры"""
        if not self.active_filters:
            return

        print(f"  reapply_all_filters called with: {self.active_filters}")

        # Показываем все элементы
        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        # Применяем все активные фильтры
        visible_items = set(self.tree.get_children(''))

        for col_id, values in self.active_filters.items():
            try:
                col_index = list(self.tree["columns"]).index(col_id)
                items_to_hide = set()

                for item_id in visible_items:
                    try:
                        item_values = self.tree.item(item_id)["values"]
                        value = str(item_values[col_index])

                        if value not in values:
                            items_to_hide.add(item_id)
                    except:
                        pass

                for item_id in items_to_hide:
                    self.tree.detach(item_id)

                visible_items -= items_to_hide
            except:
                pass

        self.update_column_headers()

    def show_filter_menu(self, event, column_id):
        """Показать меню фильтра для столбца"""
        try:
            column_index = list(self.tree["columns"]).index(column_id)

            # Собираем уникальные значения
            all_unique_values = set()
            visible_unique_values = set()

            visible_items = list(self.tree.get_children(''))

            if not hasattr(self, '_all_item_cache'):
                self._all_item_cache = set()

            for item_id in visible_items:
                self._all_item_cache.add(item_id)
                values = self.tree.item(item_id)["values"]
                value = values[column_index]
                visible_unique_values.add(str(value))
                all_unique_values.add(str(value))

            for item_id in self._all_item_cache:
                if item_id not in visible_items:
                    try:
                        values = self.tree.item(item_id)["values"]
                        if values:
                            value = values[column_index]
                            all_unique_values.add(str(value))
                    except:
                        pass

            # Определяем выбранные значения
            if column_id in self.active_filters:
                currently_selected = set(self.active_filters[column_id])
            else:
                currently_selected = all_unique_values.copy()

            # Сортировка
            selected_visible = sorted(list(currently_selected & visible_unique_values))
            selected_hidden = sorted(list(currently_selected - visible_unique_values))
            unselected = sorted(list(all_unique_values - currently_selected))

            unique_values = selected_visible + selected_hidden + unselected

            # Создаём окно фильтра
            filter_window = tk.Toplevel(self.tree)
            filter_window.title(f"Фильтр: {column_id}")
            filter_window.geometry("320x600")
            filter_window.configure(bg='#ecf0f1')
            filter_window.transient(self.tree)
            filter_window.grab_set()

            x = event.x_root
            y = event.y_root + 20
            filter_window.geometry(f"+{x}+{y}")

            # Заголовок
            header_frame = tk.Frame(filter_window, bg='#3498db')
            header_frame.pack(fill=tk.X)
            tk.Label(header_frame, text=f"Фильтр: {column_id}",
                     font=("Arial", 12, "bold"), bg='#3498db', fg='white', pady=10).pack()

            # Кнопки сортировки
            sort_frame = tk.Frame(filter_window, bg='#ecf0f1')
            sort_frame.pack(fill=tk.X, padx=10, pady=10)
            tk.Label(sort_frame, text="Сортировка:", font=("Arial", 10, "bold"), bg='#ecf0f1').pack(anchor='w',
                                                                                                    pady=(0, 5))
            tk.Button(sort_frame, text="▲ По возрастанию (A→Z, 0→9)",
                      command=lambda: self.apply_sort(column_id, 'asc', filter_window),
                      bg='#3498db', fg='white', font=("Arial", 9), relief=tk.RAISED).pack(fill=tk.X, pady=2)
            tk.Button(sort_frame, text="▼ По убыванию (Z→A, 9→0)",
                      command=lambda: self.apply_sort(column_id, 'desc', filter_window),
                      bg='#3498db', fg='white', font=("Arial", 9), relief=tk.RAISED).pack(fill=tk.X, pady=2)

            tk.Frame(filter_window, height=2, bg='#95a5a6').pack(fill=tk.X, pady=5)

            # ПОЛЕ ПОИСКА
            search_frame = tk.Frame(filter_window, bg='#ecf0f1')
            search_frame.pack(fill=tk.X, padx=10, pady=5)
            tk.Label(search_frame, text="🔍 Поиск:", font=("Arial", 10, "bold"), bg='#ecf0f1').pack(side=tk.LEFT, padx=5)

            search_var = tk.StringVar()
            search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Arial", 10), width=20)
            search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

            def clear_search():
                search_var.set("")
                search_entry.focus_set()

            tk.Button(search_frame, text="✖", font=("Arial", 8), bg='#e74c3c', fg='white',
                      command=clear_search, width=3).pack(side=tk.LEFT, padx=2)

            tk.Frame(filter_window, height=2, bg='#95a5a6').pack(fill=tk.X, pady=5)

            tk.Label(filter_window, text="Фильтр по значению:",
                     font=("Arial", 10, "bold"), bg='#ecf0f1').pack(pady=(5, 5), padx=10, anchor='w')

            # Фрейм со списком
            list_frame = tk.Frame(filter_window, bg='white', relief=tk.SUNKEN, borderwidth=1)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

            scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            canvas = tk.Canvas(list_frame, bg='white', yscrollcommand=scrollbar.set, highlightthickness=0)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=canvas.yview)

            checkboxes_frame = tk.Frame(canvas, bg='white')
            canvas_window = canvas.create_window((0, 0), window=checkboxes_frame, anchor='nw')

            # Прокрутка
            def _on_mousewheel(e):
                canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

            def _bind_to_mousewheel(e):
                canvas.bind_all("<MouseWheel>", _on_mousewheel)

            def _unbind_from_mousewheel(e):
                canvas.unbind_all("<MouseWheel>")

            canvas.bind('<Enter>', _bind_to_mousewheel)
            canvas.bind('<Leave>', _unbind_from_mousewheel)

            trace_id = None

            # Очистка при закрытии
            def cleanup_and_close():
                try:
                    _unbind_from_mousewheel(None)
                except:
                    pass

                try:
                    if trace_id is not None:
                        search_var.trace_remove('write', trace_id)
                except:
                    pass

                self._filter_window_open = False

                try:
                    filter_window.destroy()
                except:
                    pass

            filter_window.protocol("WM_DELETE_WINDOW", cleanup_and_close)

            # "Выбрать всё"
            all_selected = (len(currently_selected.intersection(unique_values)) == len(unique_values))
            select_all_var = tk.BooleanVar(value=all_selected)
            checkbox_vars = {}

            def toggle_all():
                state = select_all_var.get()
                search_text = search_var.get().lower().strip()
                for value, var in checkbox_vars.items():
                    if not search_text or search_text in value.lower():
                        var.set(state)

            select_all_frame = tk.Frame(checkboxes_frame, bg='#e8f4f8')
            select_all_frame.pack(fill=tk.X, pady=2)
            tk.Checkbutton(select_all_frame, text="✓ Выбрать всё",
                           variable=select_all_var, command=toggle_all,
                           font=("Arial", 10, "bold"), bg='#e8f4f8',
                           activebackground='#d1ecf1').pack(anchor='w', padx=5, pady=5)

            tk.Frame(checkboxes_frame, height=2, bg='#95a5a6').pack(fill=tk.X, padx=5, pady=2)

            checkbox_frames = {}

            # Создаём чекбоксы
            for value in unique_values:
                is_checked = (value in currently_selected)
                is_visible = (value in visible_unique_values)

                var = tk.BooleanVar(value=is_checked)
                checkbox_vars[value] = var

                bg_color = 'white' if is_visible else '#f8f8f8'
                cb_frame = tk.Frame(checkboxes_frame, bg=bg_color)
                cb_frame.pack(fill=tk.X, padx=2, pady=1)

                display_text = f"{value} 🔒" if not is_visible else value
                cb = tk.Checkbutton(cb_frame, text=display_text, variable=var,
                                    font=("Arial", 9, "italic" if not is_visible else "normal"),
                                    bg=bg_color, fg='#888' if not is_visible else 'black',
                                    activebackground='#e0e0e0' if not is_visible else '#f0f0f0')
                cb.pack(anchor='w', padx=10, pady=2)

                checkbox_frames[value] = cb_frame

            # ФУНКЦИЯ ФИЛЬТРАЦИИ
            def filter_checkboxes(*args):
                try:
                    if not filter_window.winfo_exists():
                        return
                except:
                    return

                search_text = search_var.get().lower().strip()

                # Скрываем все
                for cb_frame in checkbox_frames.values():
                    cb_frame.pack_forget()

                # Показываем нужные
                for value in unique_values:
                    if value not in checkbox_vars:
                        continue

                    var = checkbox_vars[value]
                    cb_frame = checkbox_frames[value]

                    if not search_text:
                        cb_frame.pack(fill=tk.X, padx=2, pady=1)
                    elif search_text in value.lower():
                        cb_frame.pack(fill=tk.X, padx=2, pady=1)
                        var.set(True)
                    else:
                        var.set(False)

                # Обновляем "Выбрать всё"
                if search_text:
                    checked = sum(1 for v in checkbox_vars.values() if v.get())
                    select_all_var.set(checked > 0)
                else:
                    checked = sum(1 for v in checkbox_vars.values() if v.get())
                    select_all_var.set(checked == len(checkbox_vars))

                # Обновляем canvas
                try:
                    checkboxes_frame.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                except:
                    pass

            trace_id = search_var.trace_add('write', filter_checkboxes)
            search_entry.focus_set()

            def on_frame_configure(event=None):
                try:
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    canvas.itemconfig(canvas_window, width=canvas.winfo_width())
                except:
                    pass

            checkboxes_frame.bind("<Configure>", on_frame_configure)
            canvas.bind("<Configure>", on_frame_configure)

            # Функции для кнопок
            def apply_value_filter():
                selected_values = {value for value, var in checkbox_vars.items() if var.get()}
                if not selected_values:
                    messagebox.showwarning("Предупреждение", "Выберите хотя бы одно значение!")
                    return
                cleanup_and_close()
                self.apply_filter(column_id, selected_values, None)

            def clear_filter():
                if column_id in self.active_filters:
                    del self.active_filters[column_id]
                self.update_column_headers()
                cleanup_and_close()
                self.refresh_callback()

            # Кнопки
            buttons_frame = tk.Frame(filter_window, bg='#ecf0f1')
            buttons_frame.pack(fill=tk.X, padx=10, pady=10)

            tk.Button(buttons_frame, text="✓ Применить фильтр", command=apply_value_filter,
                      bg='#27ae60', fg='white', font=("Arial", 10, "bold"), relief=tk.RAISED, borderwidth=2).pack(
                side=tk.LEFT, padx=5, expand=True, fill=tk.X)

            tk.Button(buttons_frame, text="✗ Сбросить фильтр", command=clear_filter,
                      bg='#e74c3c', fg='white', font=("Arial", 10)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

            tk.Button(buttons_frame, text="Отмена", command=cleanup_and_close,
                      bg='#95a5a6', fg='white', font=("Arial", 10)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        except Exception as e:
            print(f"❌ ОШИБКА: {e}")
            import traceback
            traceback.print_exc()
            self._filter_window_open = False

    def apply_sort(self, column_id, direction, window):
        """Применить сортировку"""
        column_index = list(self.tree["columns"]).index(column_id)

        items = [(self.tree.item(item_id)["values"], item_id) for item_id in self.tree.get_children()]

        try:
            items.sort(key=lambda x: float(str(x[0][column_index]).replace(',', '.')),
                       reverse=(direction == 'desc'))
        except:
            items.sort(key=lambda x: str(x[0][column_index]).lower(),
                       reverse=(direction == 'desc'))

        for index, (values, item_id) in enumerate(items):
            self.tree.move(item_id, '', index)

        window.destroy()
        self._filter_window_open = False

    def apply_filter(self, column_id, selected_values, sort_order):
        """Применить фильтр"""
        print(f"\n🔍 DEBUG apply_filter:")
        print(f"  column_id: {column_id}")
        print(f"  selected_values: {selected_values}")

        self.active_filters[column_id] = selected_values
        print(f"  active_filters: {self.active_filters}")

        column_index = list(self.tree["columns"]).index(column_id)
        print(f"  column_index: {column_index}")

        # Показываем все из кэша
        print(f"  _all_item_cache size: {len(self._all_item_cache)}")
        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        # Применяем ВСЕ активные фильтры
        visible_items = set(self.tree.get_children(''))
        print(f"  visible_items before filter: {len(visible_items)}")

        for col_id, values in self.active_filters.items():
            col_index = list(self.tree["columns"]).index(col_id)
            print(f"  Processing filter for {col_id} (index {col_index})")

            items_to_hide = set()

            for item_id in visible_items:
                try:
                    item_values = self.tree.item(item_id)["values"]
                    value = str(item_values[col_index])

                    if value not in values:
                        items_to_hide.add(item_id)
                        print(f"    Hiding item {item_id}: value='{value}' not in {values}")
                except Exception as e:
                    print(f"    Error processing item {item_id}: {e}")
                    pass

            print(f"  items_to_hide: {len(items_to_hide)}")
            for item_id in items_to_hide:
                self.tree.detach(item_id)

            visible_items -= items_to_hide

        print(f"  visible_items after filter: {len(visible_items)}")
        self.update_column_headers()

        if self.refresh_callback:
            print(f"  Calling refresh_callback")
            self.refresh_callback()

    def update_column_headers(self):
        """Обновить заголовки колонок (добавить/убрать индикатор фильтра)"""
        for col in self.tree["columns"]:
            current_text = col.replace(" 🔽", "")

            if col in self.active_filters or current_text in self.active_filters:
                new_text = f"{current_text} 🔽"
            else:
                new_text = current_text

            self.tree.heading(col, text=new_text)

    def clear_all_filters(self):
        """очистить все фильтры"""
        self.active_filters = {}

        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        self.update_column_headers()

        if self.refresh_callback:
            self.refresh_callback()


# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА "ТИПЫ КОМПЛЕКТУЮЩИХ"
# ─────────────────────────────────────────────────────────────

class ComponentTypesTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self._build_ui()

    def _build_ui(self):
        # Панель кнопок
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=6, pady=(6, 2))

        ttk.Button(btn_bar, text="➕ Добавить тип", command=self.add_type).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="✏️ Редактировать", command=self.edit_type).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🗑️ Удалить", command=self.delete_type).pack(side="left", padx=2)

        # Таблица с прокруткой
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        self.tree = ttk.Treeview(tree_frame, columns=["название", "описание"], show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.heading("название", text="Название типа", anchor="w")
        self.tree.heading("описание", text="Описание", anchor="w")
        self.tree.column("название", width=150, minwidth=100, anchor="w")
        self.tree.column("описание", width=300, minwidth=100, anchor="w")

        self.tree.bind("<Double-1>", lambda e: self.open_details())

        # Строка состояния
        self._status_var = tk.StringVar()
        ttk.Label(self, textvariable=self._status_var, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4))

    def refresh(self):
        """Перерисовать таблицу типов."""
        self.tree.delete(*self.tree.get_children())
        for type_item in self.app.component_types:
            self.tree.insert("", "end", values=[
                type_item.get("название", ""),
                type_item.get("описание", "")
            ])
        self._status_var.set(f"Типов: {len(self.app.component_types)}")

    def open_details(self):
        """Открыть вкладку с деталями типа."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите тип для просмотра.")
            return

        row_values = self.tree.item(sel[0])["values"]
        type_name = row_values[0]

        # Найти type_id
        type_item = next((t for t in self.app.component_types if t.get("название") == type_name), None)
        if type_item is None:
            return

        self.app.selected_type = type_item
        self.app.notebook.select(1)  # Переключиться на вкладку "Детали типа"
        self.app.tab_details.refresh()

    def add_type(self):
        dlg = ComponentTypeDialog(self.winfo_toplevel(), "Добавить тип комплектующего")
        self.wait_window(dlg)
        if dlg.result is None:
            return

        new_type = dlg.result
        new_type["id"] = str(next_type_id(self.app.component_types))
        self.app.component_types.append(new_type)

        self.app.add_log(
            operation="Добавление типа",
            component=new_type.get("название", ""),
            change=f"Добавлен новый тип: {new_type.get('название')}",
        )
        self.app.auto_save()
        self.refresh()

    def edit_type(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите тип для редактирования.")
            return

        row_values = self.tree.item(sel[0])["values"]
        type_name = row_values[0]

        type_item = next((t for t in self.app.component_types if t.get("название") == type_name), None)
        if type_item is None:
            return

        old_type = copy.deepcopy(type_item)
        dlg = ComponentTypeDialog(self.winfo_toplevel(), "Редактировать тип", type_item)
        self.wait_window(dlg)
        if dlg.result is None:
            return

        type_item.update(dlg.result)

        if old_type.get("название") != dlg.result.get("название") or old_type.get("описание") != dlg.result.get(
                "описание"):
            self.app.add_log(
                operation="Изменение типа",
                component=dlg.result.get("название", ""),
                change=f"Название: {old_type.get('название')} → {dlg.result.get('название')}",
            )

        self.app.auto_save()
        self.refresh()

    def delete_type(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите тип для удаления.")
            return

        row_values = self.tree.item(sel[0])["values"]
        type_name = row_values[0]

        type_item = next((t for t in self.app.component_types if t.get("название") == type_name), None)
        if type_item is None:
            return

        # Проверяем есть ли комплектующие этого типа
        components_count = len([c for c in self.app.components if c.get("type_id") == type_item.get("id")])

        if components_count > 0:
            messagebox.showwarning("Внимание",
                                   f"Не можно удалить тип с {components_count} комплектующими.\nСначала удалите все комплектующие этого типа.")
            return

        if not messagebox.askyesno("Удаление", f"Удалить тип «{type_name}»?"):
            return

        self.app.component_types = [t for t in self.app.component_types if t.get("id") != type_item.get("id")]
        self.app.add_log(
            operation="Удаление типа",
            component=type_name,
            change="Тип комплектующего удалён.",
        )
        self.app.auto_save()
        self.refresh()


# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА "ДЕТАЛИ ТИПА КОМПЛЕКТУЮЩИХ"
# ─────────────────────────────────────────────────────────────

class ComponentDetailsTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self._filter: ExcelStyleFilter | None = None
        self._build_ui()

    def _build_ui(self):
        # Панель кнопок
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=6, pady=(6, 2))

        ttk.Button(btn_bar, text="◀️ Назад", command=self.go_back).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="➕ Добавить", command=self.add_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="✏️ Редактировать", command=self.edit_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🗑️ Удалить", command=self.delete_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🔄 Сбросить фильтры", command=self.reset_filters).pack(side="left", padx=8)

        self._filter_status = ttk.Label(btn_bar, text="", foreground="blue")
        self._filter_status.pack(side="left", padx=4)

        self._type_label = ttk.Label(btn_bar, text="", font=("", 10, "bold"), foreground="darkblue")
        self._type_label.pack(side="right", padx=10)

        # Таблица с прокруткой
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        self.tree = ttk.Treeview(tree_frame, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.bind("<Double-1>", lambda e: self.edit_item())

        # Строка состояния
        self._status_var = tk.StringVar()
        ttk.Label(self, textvariable=self._status_var, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4))

    def _setup_columns(self, extra_keys: list[str]):
        """Настроить столбцы таблицы с учётом доп. параметров типа."""
        # 🆕 Используем параметры из типа
        type_params = []
        if hasattr(self.app, 'selected_type') and self.app.selected_type:
            type_params = self.app.selected_type.get("параметры", [])

        cols = list(STANDARD_COLUMNS) + type_params
        self.tree.configure(columns=cols)
        for c in cols:
            header = STANDARD_HEADERS.get(c, c)
            self.tree.heading(c, text=header, anchor="w")
            width = 120 if c not in ("id", "тип") else (40 if c == "id" else 90)
            self.tree.column(c, width=width, minwidth=50, anchor="w")

        # Устанавливаем фильтр
        if self._filter is None:
            self._filter = ExcelStyleFilter(self.tree, self.refresh)
        else:
            print(f"🔧 Filter already exists, keeping filters: {self._filter.active_filters}")

    def _collect_extra_keys(self) -> list[str]:
        """Получить параметры из выбранного типа."""
        if hasattr(self.app, 'selected_type') and self.app.selected_type:
            return self.app.selected_type.get("параметры", [])
        return []

    def refresh(self):
        """Перерисовать таблицу деталей типа."""
        if not hasattr(self.app, 'selected_type') or self.app.selected_type is None:
            self._type_label.config(text="Выберите тип")
            self.tree.delete(*self.tree.get_children())
            return

        type_id = self.app.selected_type.get("id")
        type_name = self.app.selected_type.get("название", "")

        extra_keys = self._collect_extra_keys()
        self._setup_columns(extra_keys)

        # Фильтруем комплектующие по type_id
        filtered_components = [c for c in self.app.components if c.get("type_id") == type_id]

        all_rows = []
        for item in filtered_components:
            row = [item.get(c, "") for c in STANDARD_COLUMNS]
            for k in extra_keys:
                row.append(item.get("доп_параметры", {}).get(k, ""))
            all_rows.append(tuple(row))

        # Очищаем таблицу
        self.tree.delete(*self.tree.get_children())

        # Заполняем таблицу всеми строками
        for row in all_rows:
            self.tree.insert("", "end", values=row)

        # Инициализируем кэш фильтра
        if self._filter:
            if not hasattr(self._filter, '_all_item_cache'):
                self._filter._all_item_cache = set()
            self._filter._all_item_cache = set(self.tree.get_children(''))

            # Переприменяем фильтры
            if self._filter.active_filters:
                self._filter.reapply_all_filters()
                visible_count = len(self.tree.get_children(''))
            else:
                visible_count = len(all_rows)
        else:
            visible_count = len(all_rows)

        self._type_label.config(text=f"Тип: {type_name}")
        self._status_var.set(f"Показано: {visible_count} из {len(filtered_components)}")

        if self._filter and self._filter.active_filters:
            self._filter_status.config(text="⚠ Активны фильтры")
        else:
            self._filter_status.config(text="")

    def go_back(self):
        """Вернуться на вкладку типов."""
        self.app.notebook.select(0)
        self.app.tab_types.refresh()

    def reset_filters(self):
        if self._filter:
            self._filter.clear_all_filters()
        self.refresh()

    def add_item(self):
        if not hasattr(self.app, 'selected_type') or self.app.selected_type is None:
            messagebox.showinfo("Внимание", "Выберите тип комплектующего сначала.")
            return

        dlg = ComponentDialog(self.winfo_toplevel(), "Добавить комплектующее", app=self.app)  # 🆕 app=self.app
        self.wait_window(dlg)
        if dlg.result is None:
            return
        new_item = dlg.result
        new_item["id"] = str(next_id(self.app.components))
        new_item["type_id"] = self.app.selected_type.get("id")
        self.app.components.append(new_item)
        self.app.add_log(
            operation="Добавление",
            component=component_label(new_item),
            change=f"Добавлено новое комплектующее: {component_label(new_item)}",
        )
        self.app.auto_save()
        self.refresh()

    def edit_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите строку для редактирования.")
            return
        row_values = self.tree.item(sel[0])["values"]
        item_id = str(row_values[0])
        item = next((c for c in self.app.components if str(c.get("id")) == item_id), None)
        if item is None:
            return

        old_item = copy.deepcopy(item)
        dlg = ComponentDialog(self.winfo_toplevel(), "Редактировать комплектующее", item, app=self.app)  # 🆕 app=self.app
        self.wait_window(dlg)
        if dlg.result is None:
            return

        change_desc = diff_items(old_item, dlg.result)
        item.update(dlg.result)
        if change_desc != "Без изменений":
            self.app.add_log(
                operation="Изменение",
                component=component_label(item),
                change=change_desc,
            )
        self.app.auto_save()
        self.refresh()

    def delete_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите строку для удаления.")
            return
        row_values = self.tree.item(sel[0])["values"]
        item_id = str(row_values[0])
        item = next((c for c in self.app.components if str(c.get("id")) == item_id), None)
        if item is None:
            return
        label = component_label(item)
        if not messagebox.askyesno("Удаление", f"Удалить «{label}»?"):
            return
        self.app.components = [c for c in self.app.components if str(c.get("id")) != item_id]
        self.app.add_log(
            operation="Удаление",
            component=label,
            change="Комплектующее удалено.",
        )
        self.app.auto_save()
        self.refresh()

# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА "ЛОГ"
# ─────────────────────────────────────────────────────────────

class LogTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self._build_ui()

    def _build_ui(self):
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Button(btn_bar, text="💬 Добавить комментарий", command=self._add_comment).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🔄 Обновить", command=self.refresh).pack(side="left", padx=2)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        cols = list(LOG_COLUMNS)
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        for c in cols:
            header = LOG_HEADERS.get(c, c)
            self.tree.heading(c, text=header, anchor="w",
                              command=lambda col=c: self._sort_by(col))
            widths = {"дата_время": 140, "операция": 90, "комплектующее": 150,
                      "изменение": 300, "комментарий": 200}
            self.tree.column(c, width=widths.get(c, 120), minwidth=60, anchor="w")

        self.tree.bind("<Double-1>", lambda e: self._add_comment())
        self._sort_col = "дата_время"
        self._sort_asc = False

        self._status_var = tk.StringVar()
        ttk.Label(self, textvariable=self._status_var, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4))

    def _sort_by(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self.refresh()

    def refresh(self):
        entries = sorted(
            self.app.log_entries,
            key=lambda e: e.get(self._sort_col, ""),
            reverse=not self._sort_asc,
        )
        self.tree.delete(*self.tree.get_children())
        for e in entries:
            self.tree.insert("", "end", values=[e.get(c, "") for c in LOG_COLUMNS])
        self._status_var.set(f"Записей в журнале: {len(entries)}")

    def _add_comment(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите запись лога для добавления комментария.")
            return
        row_values = self.tree.item(sel[0])["values"]
        dt_val = row_values[0] if row_values else ""

        # Найти запись в логе по дате
        entry = next(
            (e for e in self.app.log_entries if e.get("дата_время") == dt_val),
            None,
        )
        if entry is None:
            return

        current = entry.get("комментарий", "")
        new_comment = simpledialog.askstring(
            "Комментарий", "Введите комментарий:",
            initialvalue=current,
            parent=self.winfo_toplevel(),
        )
        if new_comment is None:
            return
        entry["комментарий"] = new_comment
        self.app.auto_save()
        self.refresh()


# ─────────────────────────────────────────────────────────────
#  ДИАЛОГ НАСТРОЕК
# ─────────────────────────────────────────────────────────────

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, settings: dict):
        super().__init__(parent)
        self.title("Настройки")
        self.resizable(False, False)
        self.grab_set()
        self.result = None
        self._settings = dict(settings)
        self._build_ui()
        self.update_idletasks()
        w, h = 480, 160
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        frame = ttk.Frame(self, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Папка для Excel-файла:").grid(row=0, column=0, sticky="w", pady=4)
        self._path_var = tk.StringVar(
            value=self._settings.get("db_folder",
                                     os.path.dirname(os.path.abspath(__file__))))
        path_entry = ttk.Entry(frame, textvariable=self._path_var, width=40)
        path_entry.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)
        ttk.Button(frame, text="...", width=3, command=self._browse).grid(row=0, column=2, padx=4)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Файл будет сохранён как:").grid(row=1, column=0, sticky="w", pady=2)
        self._file_label = ttk.Label(frame, text=DEFAULT_DB_NAME, foreground="grey")
        self._file_label.grid(row=1, column=1, columnspan=2, sticky="w", padx=4)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=3, sticky="e", pady=8)
        ttk.Button(btn_frame, text="Сохранить", command=self._ok).pack(side="right", padx=4)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side="right")

    def _browse(self):
        folder = filedialog.askdirectory(
            title="Выберите папку для Excel",
            initialdir=self._path_var.get(),
            parent=self,
        )
        if folder:
            self._path_var.set(folder)

    def _ok(self):
        self._settings["db_folder"] = self._path_var.get()
        self.result = self._settings
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  ГЛАВНОЕ ПРИЛОЖЕНИЕ
# ─────────────────────────────────────────────────────────────

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("VitaKa — Учёт комплектующих")
        self.minsize(800, 500)

        self.settings = load_settings()
        self._restore_geometry()

        # Данные
        db_path = get_db_path(self.settings)
        initialize_db(db_path)
        self.component_types = load_types(db_path)
        self.components = load_components(db_path)
        self.log_entries = load_log(db_path)
        self.selected_type = None

        self._build_menu()
        self._build_ui()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.bind("<Configure>", self._on_resize)

    # ── интерфейс ─────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self)
        self.configure(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Настройки…", command=self._open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Открыть папку с Excel", command=self._open_db_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self._on_close)

    def _build_ui(self):
        # 🆕 СОЗДАЁМ _status_var СНАЧАЛА
        status_bar = ttk.Frame(self, relief="sunken")
        status_bar.pack(fill="x", side="bottom")
        self._status_var = tk.StringVar(value="Готово")
        ttk.Label(status_bar, textvariable=self._status_var, anchor="w",
                  padding=(6, 2)).pack(side="left")
        self._db_path_label = ttk.Label(
            status_bar, text=get_db_path(self.settings), anchor="e",
            foreground="grey", padding=(6, 2))
        self._db_path_label.pack(side="right")

        # ПОТОМ создаём вкладки
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=4, pady=4)

        self.tab_types = ComponentTypesTab(self.notebook, self)
        self.notebook.add(self.tab_types, text="Типы комплектующих")

        self.tab_details = ComponentDetailsTab(self.notebook, self)
        self.notebook.add(self.tab_details, text="Детали типа")

        self.tab_log = LogTab(self.notebook, self)
        self.notebook.add(self.tab_log, text="Журнал изменений")

        self.tab_types.refresh()
        self.tab_log.refresh()

    # ── лог ───────────────────────────────────────────────

    def add_log(self, operation: str, component: str, change: str, comment: str = ""):
        entry = {
            "дата_время": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "операция": operation,
            "комплектующее": component,
            "изменение": change,
            "комментарий": comment,
        }
        self.log_entries.append(entry)
        if hasattr(self, "tab_log"):
            self.tab_log.refresh()

    # ── сохранение ────────────────────────────────────────

    def auto_save(self):
        db_path = get_db_path(self.settings)
        save_all_with_types(db_path, self.component_types, self.components, self.log_entries)
        self._status_var.set(f"Сохранено: {datetime.now().strftime('%H:%M:%S')}")
        self._db_path_label.config(text=db_path)

    # ── настройки ─────────────────────────────────────────

    def _open_settings(self):
        dlg = SettingsDialog(self, self.settings)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        old_path = get_db_path(self.settings)
        self.settings = dlg.result
        save_settings(self.settings)
        new_path = get_db_path(self.settings)
        initialize_db(new_path)
        # Перезагрузить данные из нового места, если файл там уже есть
        if os.path.exists(new_path) and new_path != old_path:
            self.components = load_components(new_path)
            self.log_entries = load_log(new_path)
        else:
            # Сохранить текущие данные в новое место
            self.auto_save()
        self.tab_components.refresh()
        self.tab_log.refresh()
        self._db_path_label.config(text=new_path)

    def _open_db_folder(self):
        folder = os.path.dirname(get_db_path(self.settings))
        if os.path.exists(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        else:
            messagebox.showinfo("Папка", f"Папка не найдена:\n{folder}")

    # ── геометрия окна ────────────────────────────────────

    def _restore_geometry(self):
        geom = self.settings.get("window_geometry", "1100x650+100+50")
        try:
            self.geometry(geom)
        except Exception:
            self.geometry("1100x650")

    def _on_resize(self, event):
        # Сохраняем геометрию только для главного окна
        if event.widget is self:
            self.settings["window_geometry"] = self.geometry()

    def _on_close(self):
        save_settings(self.settings)
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  ТОЧКА ВХОДА
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = MainApp()
    app.mainloop()
